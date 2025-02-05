import openpyxl
from tkinter import *
from tkinter import messagebox
from datetime import datetime
import os
import subprocess

# Função para atualizar a planilha
def atualizar_planilha():
    patrimonio = entry_patrimonio.get()
    responsavel = entry_responsavel.get()
    setor = entry_setor.get()
    estado = estado_var.get()
    data_atualizacao = datetime.now().strftime("%d/%m/%Y")

    # Mapear os valores de estado para os nomes correspondentes
    estado_map = {
        "1": "em uso",
        "2": "estoque",
        "3": "descarte"
    }
    estado_nome = estado_map.get(estado, "")

    # Verificar se o arquivo existe
    if not os.path.exists('Inventario2023.xlsx'):
        messagebox.showerror("Erro", "Arquivo Inventario2023.xlsx não encontrado")
        return

    # Verificar permissões de leitura e escrita
    if not os.access('Inventario2023.xlsx', os.R_OK) or not os.access('Inventario2023.xlsx', os.W_OK):
        messagebox.showerror("Erro", "Permissão negada para ler ou escrever no arquivo Inventario2023.xlsx")
        return

    # Carregar a planilha
    try:
        wb = openpyxl.load_workbook('Inventario2023.xlsx')
        if 'Inventario_2024.1' not in wb.sheetnames:
            messagebox.showerror("Erro", f"Aba 'Inventario_2024.1' não encontrada. Abas disponíveis: {', '.join(wb.sheetnames)}")
            return
        sheet = wb['Inventario_2024.1']
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar a planilha: {e}")
        return

    # Procurar o patrimônio na coluna I (índice 8)
    for row in sheet.iter_rows(min_row=2, max_col=25, max_row=sheet.max_row):
        if str(row[8].value) == patrimonio:
            row[9].value = responsavel  # Coluna J (índice 9)
            row[21].value = setor       # Coluna V (índice 21)
            row[22].value = estado_nome # Coluna W (índice 22)
            row[24].value = data_atualizacao  # Coluna Y (índice 24)
            break
    else:
        messagebox.showerror("Erro", "Patrimônio não encontrado")

    # Salvar a planilha
    try:
        wb.save('Inventario2023.xlsx')
        
        # Abrir o arquivo no Excel para sincronizar com o OneDrive
        subprocess.Popen(['start', 'excel.exe', 'Inventario2023.xlsx'], shell=True)
        
        messagebox.showinfo("Sucesso", "Dados atualizados com sucesso. O Excel será aberto para sincronização.")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar a planilha: {e}")

# Criar a interface gráfica
root = Tk()
root.title("Atualizar Inventário")

Label(root, text="Patrimônio:").grid(row=0, column=0, padx=10, pady=10)
entry_patrimonio = Entry(root)
entry_patrimonio.grid(row=0, column=1, padx=10, pady=10)

Label(root, text="Responsável:").grid(row=1, column=0, padx=10, pady=10)
entry_responsavel = Entry(root)
entry_responsavel.grid(row=1, column=1, padx=10, pady=10)

Label(root, text="Setor:").grid(row=2, column=0, padx=10, pady=10)
entry_setor = Entry(root)
entry_setor.grid(row=2, column=1, padx=10, pady=10)

Label(root, text="Estado:").grid(row=3, column=0, padx=10, pady=10)
estado_var = StringVar(value="1")
Radiobutton(root, text="Em uso", variable=estado_var, value="1").grid(row=3, column=1, padx=10, pady=10, sticky=W)
Radiobutton(root, text="Estoque", variable=estado_var, value="2").grid(row=3, column=2, padx=10, pady=10, sticky=W)
Radiobutton(root, text="Descarte", variable=estado_var, value="3").grid(row=3, column=3, padx=10, pady=10, sticky=W)

Button(root, text="Atualizar", command=atualizar_planilha).grid(row=4, column=0, columnspan=4, pady=20)

root.mainloop()
