import openpyxl
from tkinter import *
from tkinter import messagebox
from datetime import datetime
import os
import subprocess
import win32com.client
import time

# Função para atualizar a planilha
def atualizar_planilha():
    patrimonio = entry_patrimonio.get()
    responsavel = entry_responsavel.get()
    setor = entry_setor.get()
    estado = estado_var.get()
    data_atualizacao = datetime.now().strftime("%d/%m/%Y")

    # Mapear os valores de estado para os nomes correspondentes
    estado_map = {
        "Em uso": "em uso",
        "Estoque": "estoque",
        "Descarte": "descarte"
    }
    estado_nome = estado_map.get(estado, "")

    # Caminho absoluto do arquivo
    file_path = os.path.abspath('Inventario2023.xlsx')

    # Verificar se o arquivo existe
    if not os.path.exists(file_path):
        messagebox.showerror("Erro", f"Arquivo {file_path} não encontrado")
        return

    # Verificar permissões de leitura e escrita
    if not os.access(file_path, os.R_OK) or not os.access(file_path, os.W_OK):
        messagebox.showerror("Erro", f"Permissão negada para ler ou escrever no arquivo {file_path}")
        return

    # Carregar a planilha
    try:
        wb = openpyxl.load_workbook(file_path)
        if 'Inventario_2024.1' not in wb.sheetnames:
            messagebox.showerror("Erro", f"Aba 'Inventario_2024.1' não encontrada. Abas disponíveis: {', '.join(wb.sheetnames)}")
            return
        sheet = wb['Inventario_2024.1']
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar a planilha: {e}")
        return

    # Procurar o patrimônio na coluna I (índice 8)
    for row in sheet.iter_rows(min_row=2, max_col=25, max_row=sheet.max_row):
        if str(row[8].value) == patrimonio:
            row[9].value = responsavel  # Coluna J (índice 9)
            row[21].value = setor       # Coluna V (índice 21)
            row[22].value = estado_nome # Coluna W (índice 22)
            row[24].value = data_atualizacao  # Coluna Y (índice 24)
            break
    else:
        messagebox.showerror("Erro", "Patrimônio não encontrado")

    # Salvar a planilha usando o Excel
    try:
        wb.save(file_path)
        
        # Abrir o arquivo no Excel para sincronizar com o OneDrive sem modo de leitura
        excel = win32com.client.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(file_path)
        
        # Deixar o Excel aberto por 1 minuto antes de fechar para sincronizar
        time.sleep(60)
        
        workbook.Save()
        workbook.Close()
        excel.Quit()
        
        messagebox.showinfo("Sucesso", "Dados atualizados com sucesso. O Excel foi aberto para sincronização.")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar a planilha: {e}")

# Criar a interface gráfica
root = Tk()
root.title("Atualizar Inventário")

# Adicionando a imagem
logo = PhotoImage(file="unifanor_logo.png")
logo_label = Label(root, image=logo)
logo_label.grid(row=0, column=0, columnspan=4, padx=10, pady=10)

Label(root, text="Patrimônio:").grid(row=1, column=0, padx=10, pady=10)
entry_patrimonio = Entry(root)
entry_patrimonio.grid(row=1, column=1, padx=10, pady=10)

Label(root, text="Responsável:").grid(row=2, column=0, padx=10, pady=10)
entry_responsavel = Entry(root)
entry_responsavel.grid(row=2, column=1, padx=10, pady=10)

Label(root, text="Setor:").grid(row=3, column=0, padx=10, pady=10)
entry_setor = Entry(root)
entry_setor.grid(row=3, column=1, padx=10, pady=10)

Label(root, text="Estado:").grid(row=4, column=0, padx=10, pady=10)
estado_var = StringVar(root)
estado_var.set("Em uso")  # valor padrão
estado_menu = OptionMenu(root, estado_var, "Em uso", "Estoque", "Descarte")
estado_menu.grid(row=4, column=1, padx=10, pady=10)

Button(root, text="Atualizar", command=atualizar_planilha).grid(row=5, column=0, columnspan=4, pady=20)

root.mainloop()
